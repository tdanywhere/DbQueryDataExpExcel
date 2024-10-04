# 
# Description: 
# This script queries the sandbox table in the Oracle database, displays the result in a table and
# adds the fetched data to an Excel file.
#
# Use >pyinstaller DbQueryDataExpExcel.py to create an executable (e.g. for use with Jenkins).
from datetime import date
import pandas as pd
import yaml
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Series, Reference

from DbConnectJdbc import OracleDB
#from DbConnectTns import OracleDB



# Main function
print("Start")

# Load database connection info from YAML file
with open('./yaml/db_connection.yaml', 'r') as file:
    db_config = yaml.safe_load(file)

user = db_config['db_login'][db_config['login']]['user']
password = db_config['db_login'][db_config['login']]['password']
dsn = db_config['db_login'][db_config['login']]['dsn']

# Database connection string
db = OracleDB(username=user, password=password, dsn=dsn)
connection = db.getConnection()


#
# Using OpenPyXl
#
excel_out_path_file = './output/out_DbQuery.xlsx'
excel_in_path_file = './input/DbQuery_template.xlsx'
sheet_page_raw = 'Page raw'
sheet_task_raw = 'Task raw'
sheet_cockpit = 'Cockpit'
sheet_pivots = 'Pivots'

print("Add data to new Sheet. Using OpenPyXl")
wb = openpyxl.load_workbook(excel_in_path_file)

del wb[sheet_page_raw]
del wb[sheet_task_raw]

wb.create_sheet(index=2, title=sheet_page_raw)
wb.create_sheet(index=3, title=sheet_task_raw)

########################
# Page HEADER
########################

# Query the table
query = db_config['sql_query'][db_config['sql_head']]

# Read the result into a DataFrame
df = pd.read_sql(query, con=connection)
 
# Display the result as table on standard output.
print(df)

wb.active = wb[sheet_page_raw]
ws = wb.active

# Create List and append to Head Sheet.
tree_data = df.values.tolist()
# Add Headline to Excel Sheet.
ws.append(["ID", "PAGE_ID", "PAGE_NAME", "PAGE_MODE", "STATUS", "PRIORITY", "START_DATE", "FINISH_DATE"])

# Append Rows of List to Sheet.
for row in tree_data:
    ws.append(row)

wb.save(excel_out_path_file)

# Formatting Cells.
ft = Font(bold=True)
for row in ws["A1:F1"]:
    for cell in row:
        cell.font = ft

wb.save(excel_out_path_file)

########################
# Page Task
########################

# Query the table
query = db_config['sql_query'][db_config['sql_task']]

# Read the result into a DataFrame
df = pd.read_sql(query, con=connection)
 
# Display the result as table on standard output.
print(df)

wb.active = wb[sheet_task_raw]
ws = wb.active

# Create List and append to Task Sheet.
tree_data = df.values.tolist()
# Add Headline to Excel Sheet.
ws.append(["ID", "PAGE_ID", "STATUS", "PRIORITY", "START_DATE", "FINISH_DATE", "DESCRIPTION", "CLUSTER_NAME", "DEVELOPER", "TESTER"])

# Append Rows of List to Sheet.
for row in tree_data:
    ws.append(row)

wb.save(excel_out_path_file)

# Formatting Cells.
ft = Font(bold=True)
for row in ws["A1:J1"]:
    for cell in row:
        cell.font = ft

wb.save(excel_out_path_file)

########################
# Pivots
########################
wb.active = wb[sheet_pivots]
ws = wb.active

pivot = ws._pivots[0] # any will do as they share the same cache
pivot.cache.refreshOnLoad = True

########################
# Cockpit
########################
wb.active = wb[sheet_cockpit]
ws = wb.active

ws['F1'] = date.today().strftime("%d.%m.%Y")
wb.save(excel_out_path_file)

#################################################
# Finish process and clean up.
#################################################
wb.active = wb[sheet_page_raw]
ws = wb.active
ws.sheet_state = 'hidden'

wb.active = wb[sheet_task_raw]
ws = wb.active
ws.sheet_state = 'hidden'

# Set active Sheet to Cockpit.
wb.active = wb[sheet_cockpit]
wb.save(excel_out_path_file)

# Close the connection
db.disconnect()

print("End")
# End of file